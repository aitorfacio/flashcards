import argparse
import math
from pathlib import Path

from openpyxl import load_workbook
from fpdf import FPDF
from tqdm import tqdm
from session_management import save_word_in_session, get_session_words



def page_format_to_dimensions(page_format, orientation='P'):
    if page_format == 'A4':
        if orientation == 'P':
            return 297, 210
        elif orientation == 'L':
            return 210, 297


def create_pdf_page(pdf, text_values=[], num_rows=3, num_cols=3,
                    page_format='A4',
                    orientation='L', default_text=None, dashed_lines=False):
    height, width = page_format_to_dimensions(page_format, orientation)
    #border_x = 10
    #border_y = 0
    pdf.add_page()
    rect_height = (height - pdf.get_y()*2) / num_rows
    rect_width = (width - pdf.get_x()*2) / num_cols
    #print(f"rect_height: {rect_height}")
    #print(f"rect_width: {rect_width}")
    pdf.set_font('Arial', 'B', 16)
    #print(f"x: {pdf.get_x()}, y: {pdf.get_y()}")
    index = 0
    for row in range(num_rows):
        #print("Incrementing Y")
        pdf.set_y(pdf.get_y())
        for col in range(num_cols):
            #print(f"x: {pdf.get_x()}, y: {pdf.get_y()}")
            line = 0 if col < num_cols - 1 else 1
            try:
                text = text_values[index]
            except IndexError:
                if default_text is None:
                    text = f"a_{row+1}{col+1}"
                else:
                    text = default_text
            pdf.cell(rect_width, rect_height, text, 0, line, 'C')
            index += 1
        # print dashed lines
        if dashed_lines:
            # print two vertical dashed line in a third of the width
            # the coordinates are as folows:
            # x1 the x-coordinate of the start of the line, a third of the width from the left side of the page
            # y1 the y-coordinate of the start of the line, the top of the page
            # x2 the x-coordinate of the end of the line, a third of the width from the left side of the page
            # y2 the y-coordinate of the end of the line, the bottom of the page
            pdf.dashed_line(width / 3, 0, width / 3, height)
            # print a horizontal dashed line in a third of the height
            # the coordinates are as folows:
            # x1 the x-coordinate of the start of the line, the left side of the page
            # y1 the y-coordinate of the start of the line, a third of the height from the top of the page
            # x2 the x-coordinate of the end of the line, the right side of the page
            # y2 the y-coordinate of the end of the line, a third of the height from the top of the page
            pdf.dashed_line(0, height / 3, width, height / 3)
            # print a vertical dashed line in two thirds of the width
            # the coordinates are as folows:
            # x1 the x-coordinate of the start of the line, two thirds of the width from the left side of the page
            # y1 the y-coordinate of the start of the line, the top of the page
            # x2 the x-coordinate of the end of the line, two thirds of the width from the left side of the page
            # y2 the y-coordinate of the end of the line, the bottom of the page
            pdf.dashed_line(width * 2 / 3, 0, width * 2 / 3, height)
            # print a horizontal dashed line in a third of the height
            # the coordinates are as folows:
            # x1 the x-coordinate of the start of the line, the left side of the page
            # y1 the y-coordinate of the start of the line, two thirds of the height from the top of the page
            # x2 the x-coordinate of the end of the line, the right side of the page
            # y2 the y-coordinate of the end of the line, two thirds of the height from the top of the page
            pdf.dashed_line(0, height * 2 / 3, width, height * 2 / 3)




def extract_page_from_excel(excel_rows, page_number, page_rows, page_cols):
    page_size = page_rows * page_cols
    current_rows = excel_rows[page_number * page_size: (page_number + 1) * page_size]
    # pad the current rows with a tuple of two empty strings if they are not long enough
    current_rows += [('', '')] * (page_size - len(current_rows))
    #convert the first values in the rows to a list
    first_values = [row[0] for row in current_rows]
    #convert the second values in the rows to a list, inverting the rows
    second_values = []
    for i in range(page_rows):
        row_values = [row[1] for row in current_rows[i * page_cols: (i + 1) * page_cols]]
        row_values = row_values[::-1]
        second_values.extend(row_values)
    return first_values, second_values


def list_excel_sheets(excel_file):
    wb = load_workbook(excel_file)
    return wb.sheetnames


def create_pdf_from_excel(excel_file, num_rows=3, num_cols=3,
                          output_file='output.pdf', merge_sheets=False,
                          selected_sheets=None, session_name=None,
                          complete_pages=False):
    if merge_sheets:
        pdf = FPDF(format='A4', orientation='L')
        pdf.set_auto_page_break(False)
    # if the output_file argument is None, use the name of the excel file
    if output_file is None:
        output_file = Path(excel_file).stem + '.pdf'
    page_size = num_rows * num_cols
    wb2 = load_workbook(filename=excel_file, read_only=True)
    the_sheets = wb2.worksheets
    if selected_sheets:
        the_sheets = [the_sheets[sheet] for sheet in selected_sheets]

    output_path = Path(output_file)
    if merge_sheets:
        create_merged_pdf(num_cols, num_rows, output_file, page_size, pdf, the_sheets, session_name, complete_pages)
    else:
        create_pdfs_per_worksheet(num_cols, num_rows, output_path, page_size, the_sheets, session_name, complete_pages)


def create_merged_pdf(num_cols, num_rows, output_file, page_size, pdf, the_sheets, session_name, complete_pages):
    rows = []
    for sheet in tqdm(the_sheets):
        words_in_session = []
        if session_name:
            words_in_session = get_session_words(session_name)
        sheet_rows = [(row[0].value, row[1].value) for row in sheet.iter_rows() if row[0].value not in words_in_session]
        sheet_rows = sheet_rows[1:]  # remove the first row header
        rows.extend(sheet_rows)
    if len(rows) == 0:
        return
    num_pages = math.ceil(len(rows) / page_size)
    if complete_pages:
        num_pages = math.floor(len(rows) / page_size)
    for page_number in range(num_pages):
        first_values, second_values = extract_page_from_excel(rows, page_number, num_rows, num_cols)
        # add all the first values to the session
        if session_name:
            for value in first_values:
                save_word_in_session(session_name, value)
        create_pdf_page(pdf, text_values=first_values, default_text='', num_rows=num_rows, num_cols=num_cols,
                        dashed_lines=True)
        create_pdf_page(pdf, text_values=second_values, default_text='', num_rows=num_rows, num_cols=num_cols)
    pdf.output(output_file, 'F')


def create_pdfs_per_worksheet(num_cols, num_rows, output_path, page_size, the_sheets, session_name, complete_pages):
    for sheet in tqdm(the_sheets):
        pdf = FPDF(format='A4', orientation='L')
        pdf.set_auto_page_break(False)
        words_in_session = []
        if session_name:
            words_in_session = get_session_words(session_name)
        rows = [(row[0].value, row[1].value) for row in sheet.iter_rows() if row[0].value not in words_in_session]
        rows = rows[1:]  # remove the first row header
        # the number of pages is the number of rows divided by the page size, rounded up
        num_pages = math.ceil(len(rows) / page_size)
        if complete_pages:
            num_pages = math.floor(len(rows) / page_size)
        for page_number in range(num_pages):
            first_values, second_values = extract_page_from_excel(rows, page_number, num_rows, num_cols)
            create_pdf_page(pdf, text_values=first_values, default_text='', num_rows=num_rows, num_cols=num_cols,
                            dashed_lines=True)
            create_pdf_page(pdf, text_values=second_values, default_text='', num_rows=num_rows, num_cols=num_cols)
        # append the name of the sheet to the output file and save it using pathlib
        sheet_output_file = output_path.with_name(f"{output_path.stem}_{sheet.title}.pdf")
        pdf.output(sheet_output_file)


if __name__ == '__main__':
    # read arguments from the command line and create the pdf
    parser = argparse.ArgumentParser(description='Create a pdf from an excel file')
    parser.add_argument('-f', '--file', type=str, help='The excel file to read from', required=True)
    parser.add_argument('-r', '--rows', type=int, help='The number of rows per page', default=3)
    parser.add_argument('-c', '--cols', type=int, help='The number of columns per page', default=3)
    parser.add_argument('-o', '--output', type=str, help='The output file')
    parser.add_argument('-m', '--merge', action='store_true', help='Merge all sheets into one pdf', default=False)
    parser.add_argument('-s', '--show', action='store_true', help='Show the sheets in the excel file', default=False)
    # argument to select the sheets to merge
    parser.add_argument('-t', '--sheets', type=int, help='The sheets to merge', nargs='+')
    parser.add_argument('-d', '--database_session', type=str, help='The database session to use', default=None)
    parser.add_argument('--complete_pages', action='store_true', help='Create only complete pages', default=False)
    args = parser.parse_args()
    if not args.show:
        create_pdf_from_excel(args.file, args.rows, args.cols, args.output,
                              merge_sheets=args.merge, selected_sheets=args.sheets,
                              session_name=args.database_session,complete_pages=args.complete_pages)
    else:
        sheets = list_excel_sheets(args.file)
        # print the sheets in the excel file, one each line
        for index, sheet in enumerate(sheets):
            print(f'{index} - {sheet}')
    print("Done")
